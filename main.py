import openpyxl
from openpyxl.utils import FORMULAE
import openpyxl.styles
import numpy



wb=openpyxl.Workbook()
ws=wb.active
ws.title='test1'
print(ws.title)

A=[0.17,0.13,0.215,0.125]
B=[0.1,0.2,0.3,0.4]

list1=list([A,B])
print(list1)
# for i in range(len(list1)):
#     print(list1[i])
#     print(len(list1[i]))

# ws['A1']=29
# ws.append([2,3,5])
# ws.cell(3,2).value=90
# print(ws.cell(3,2))



## 导入
for col in range(len(list1)):
    for row in range(len(list1[col])):
        ws.cell(row+1,col+1).value=list1[col][row]


# for each_row in ws.iter_rows(min_row=1,min_col=1,max_row=4,max_col=4):
#     print(each_row[0].value)

## 查看指定列数值
# for each_col in ws.rows:
#     print(each_col[0].value)


# 各列取平均值
for col in ws.iter_cols(min_col=1,min_row=1,max_col=2,max_row=6):
    ws[col[4].coordinate]='=AVERAGE(%s:%s)'%(col[0].coordinate,col[3].coordinate)
    # ws[col[4]]表示第5行第col列
    ws[col[5].coordinate].value=ws[col[4].coordinate].value
#     此处无法将纯值赋给第六行

c=ws['B6']
print(c.value)

wb.save('0917test.xlsx')

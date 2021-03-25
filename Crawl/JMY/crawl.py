import requests
from bs4 import BeautifulSoup
import openpyxl
from fake_useragent import UserAgent
import re



def get_dara():
    url='https://poedb.tw/cn/DivinationCard'
    headers = {'User-Agent': UserAgent().random}
    response = requests.get(url, headers=headers, timeout=20)

    soup=BeautifulSoup(response.text,'lxml')
    lists=soup.find_all("tr")
    del(lists[0])

    data=[]
    for l in lists:
        result=l.get_text()
        if '堆叠数量' in result:
            (title,res1)= result.split('堆叠数量',1)

            res2=re.search('\d / \d',res1)
            if res2 !=None:
                # print(res1[1:7])    #包左不包右
                if res1[6]!='1':
                    number='堆叠数量：'+res1[1:7]
                    res3=res1[7:]
                else:
                    number='堆叠数量：'+res1[1:8]
                    res3=res1[8:]
            else:
                number=' '
                res3=res1
        else:
            title=result
            number=' '
            res3=' '

        data.append([title,number,res3])



    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = '名称'
    ws['B1'] = '堆叠数量'
    ws['C1'] = '内容'
    for each in data:
        ws.append(each)
    wb.save("爬取数据.xlsx")


get_dara()